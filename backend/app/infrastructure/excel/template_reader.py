"""Read Excel template and detect structure."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class TemplateReader:
    """Reads the Excel template and provides metadata."""

    def __init__(self, template_path: str | Path):
        self.path = Path(template_path)
        if not self.path.exists():
            raise FileNotFoundError(f"Excel template not found: {self.path}")

    def load_workbook(self) -> openpyxl.Workbook:
        """Load the workbook preserving all formatting."""
        return openpyxl.load_workbook(str(self.path))

    def get_sheet(self, wb: openpyxl.Workbook) -> Worksheet:
        """Get the 'Template' sheet."""
        target = "Template"
        if target in wb.sheetnames:
            return wb[target]
        # Fallback to active sheet
        logger.warning(
            "Sheet '%s' not found, using active sheet '%s'",
            target,
            wb.active.title,
        )
        return wb.active

    def scan_labels(self, ws: Worksheet) -> dict[str, str]:
        """Scan all cells with values to help with debugging."""
        labels: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    labels[cell.coordinate] = str(cell.value)
        return labels

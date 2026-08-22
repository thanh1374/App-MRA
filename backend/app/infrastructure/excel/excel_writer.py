"""Excel writer — fills template with analysis data, preserving formatting.

Matches actual Template.xlsx layout exactly:
- Market Segmentation: values in column D (merged D:J), rows 3-15
- SWOT: B=app_name, C=strengths, D=weakness, rows 21-25
- Threats: B=ip_copyright, C=gambling_policy, D=data_providers, rows 31-33
- Customer Personas: values in column C (merged C:F), rows 40-45
- Problem Statement: B51=user, C51=problem, D51=context
- Product Idea: B57=problem, C57=vision, D57=goal, B58=target, C58=strategy, D58=feature
"""

from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from app.domain.schemas import GeminiAnalysisResult
from app.infrastructure.excel.excel_mapper import (
    CUSTOMER_PERSONAS_MAP,
    MARKET_SEGMENTATION_MAP,
    PROBLEM_STATEMENT_MAP,
    PRODUCT_IDEA_MAP,
    SWOT_COLUMNS,
    SWOT_START_ROW,
    THREATS_COLUMNS,
    THREATS_START_ROW,
)
from app.infrastructure.excel.template_reader import TemplateReader
from app.utils.sanitize import sanitize_filename

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Writes Gemini analysis into the Excel template.

    Rules:
    - Never overwrite the original template.
    - Only change cell *values* — preserve all formatting.
    - Output: market_research_<keyword>_<timestamp>.xlsx
    """

    def __init__(self, template_path: str | Path, output_dir: str | Path):
        self._template_path = Path(template_path)
        self._output_dir = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        analysis: GeminiAnalysisResult,
        keyword: str,
    ) -> Path:
        """Write analysis to a copy of the template and return output path."""

        # Generate output filename
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_keyword = sanitize_filename(keyword)
        output_name = f"market_research_{safe_keyword}_{ts}.xlsx"
        output_path = self._output_dir / output_name

        # Copy template → output (preserve original)
        shutil.copy2(str(self._template_path), str(output_path))

        # Load the copy
        reader = TemplateReader(output_path)
        wb = reader.load_workbook()
        ws = reader.get_sheet(wb)

        # Write sections
        self._write_market_segmentation(ws, analysis)
        self._write_swot(ws, analysis)
        self._write_threats(ws, analysis)
        self._write_customer_personas(ws, analysis)
        self._write_problem_statement(ws, analysis)
        self._write_product_idea(ws, analysis)

        # Save
        wb.save(str(output_path))
        logger.info("Excel written: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Section writers
    # ------------------------------------------------------------------

    def _write_market_segmentation(
        self, ws: Worksheet, analysis: GeminiAnalysisResult
    ) -> None:
        """Write Market Segmentation into column D (merged D:J)."""
        ms = analysis.market_segmentation
        data_map = {
            "D3": ms.geographical.location,
            "D4": ms.geographical.languages,
            "D5": ms.demographic.age,
            "D6": ms.demographic.gender,
            "D7": ms.demographic.income,
            "D8": ms.behavioural.occasions,
            "D9": ms.behavioural.usage_rate,
            "D10": ms.behavioural.benefits_sought,
            "D11": ms.behavioural.loyalty,
            "D12": ms.psychographic.values,
            "D13": ms.psychographic.beliefs,
            "D14": ms.psychographic.opinion,
            "D15": ms.psychographic.interests,
        }
        for cell, value in data_map.items():
            if value:
                ws[cell] = value

    def _write_swot(self, ws: Worksheet, analysis: GeminiAnalysisResult) -> None:
        """Write SWOT entries — one row per competitor app.

        Template layout (rows 21-25):
          B = App Name, C = Strengths, D = Weakness
        """
        for idx, entry in enumerate(analysis.swot[:5]):
            row = SWOT_START_ROW + idx
            ws[f"{SWOT_COLUMNS['app_name']}{row}"] = entry.app_name
            ws[f"{SWOT_COLUMNS['strengths']}{row}"] = entry.strengths
            ws[f"{SWOT_COLUMNS['weakness']}{row}"] = entry.weakness

    def _write_threats(self, ws: Worksheet, analysis: GeminiAnalysisResult) -> None:
        """Write Threats section — ip_copyright, gambling_policy, data_providers.

        Template layout (rows 31-33):
          B = IP & Copyright, C = Gambling Policy, D = Data Providers
        These are general threats (not per-app), so we aggregate from SWOT entries.
        """
        # Collect unique threat info from all SWOT entries
        ip_items = []
        gambling_items = []
        data_items = []
        for entry in analysis.swot[:5]:
            if entry.ip_copyright:
                ip_items.append(entry.ip_copyright)
            if entry.gambling_policy:
                gambling_items.append(entry.gambling_policy)
            if entry.data_providers:
                data_items.append(entry.data_providers)

        if ip_items:
            ws[f"{THREATS_COLUMNS['ip_copyright']}{THREATS_START_ROW}"] = "\n\n".join(ip_items)
        if gambling_items:
            ws[f"{THREATS_COLUMNS['gambling_policy']}{THREATS_START_ROW}"] = "\n\n".join(gambling_items)
        if data_items:
            ws[f"{THREATS_COLUMNS['data_providers']}{THREATS_START_ROW}"] = "\n\n".join(data_items)

    def _write_customer_personas(
        self, ws: Worksheet, analysis: GeminiAnalysisResult
    ) -> None:
        """Write Customer Personas into column C (merged C:F)."""
        cp = analysis.customer_personas
        data_map = {
            "C40": cp.device,
            "C41": cp.age,
            "C42": cp.needs,
            "C43": cp.painpoint,
            "C44": cp.must_have,
            "C45": cp.emotional_state,
        }
        for cell, value in data_map.items():
            if value:
                ws[cell] = value

    def _write_problem_statement(
        self, ws: Worksheet, analysis: GeminiAnalysisResult
    ) -> None:
        """Write Problem Statement.

        Template layout (row 51):
          B51 = User, C51 = Problem, D51 = Context
        """
        ps = analysis.problem_statement
        if ps.user:
            ws["B51"] = ps.user
        if ps.problem:
            ws["C51"] = ps.problem
        if ps.context:
            ws["D51"] = ps.context

    def _write_product_idea(
        self, ws: Worksheet, analysis: GeminiAnalysisResult
    ) -> None:
        """Write Product Idea.

        Template layout:
          Row 57: B=Problem, C=Vision, D=Goal
          Row 58: B=Target Audience, C=Strategy, D=Feature
        """
        pi = analysis.product_idea
        if pi.problem:
            ws["B57"] = pi.problem
        if pi.vision:
            ws["C57"] = pi.vision
        if pi.goal:
            ws["D57"] = pi.goal
        if pi.target_audience:
            ws["B58"] = pi.target_audience
        if pi.strategy:
            ws["C58"] = pi.strategy
        if pi.feature:
            ws["D58"] = pi.feature
